import json
import logging
import os
from contextlib import asynccontextmanager
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

import httpx
from dotenv import load_dotenv
from fastapi import BackgroundTasks, Depends, FastAPI, HTTPException, Query, Request, status
from fastapi.responses import FileResponse, JSONResponse, RedirectResponse, Response
from fastapi.staticfiles import StaticFiles
from starlette.middleware.sessions import SessionMiddleware
from authlib.integrations.starlette_client import OAuth

from .db import fetch_submissions as fetch_submissions_sqlite
from .db import init_db, insert_submission
from . import databricks_client as databricks

BASE_DIR = Path(__file__).resolve().parent.parent
load_dotenv(BASE_DIR / ".env")

UPLOADS_DIR = Path(os.getenv("UPLOADS_DIR", str(BASE_DIR / "uploads")))
AUTH_CLIENT_ID = os.getenv("AUTH_CLIENT_ID", "")
AUTH_CLIENT_SECRET = os.getenv("AUTH_CLIENT_SECRET", "")
AUTH_REDIRECT_URI = os.getenv("AUTH_REDIRECT_URI", "")
AUTH_COOKIE_SECRET = os.getenv("AUTH_COOKIE_SECRET", "")
AUTH_SERVER_METADATA_URL = os.getenv("AUTH_SERVER_METADATA_URL", "")

# File upload limits (match frontend script.js)
MAX_UPLOAD_SIZE_BYTES = 25 * 1024 * 1024  # 25 MB
ALLOWED_UPLOAD_EXTENSIONS = frozenset({".png", ".jpg", ".jpeg", ".pdf"})

logger = logging.getLogger(__name__)


def _is_sso_configured() -> bool:
    return bool(AUTH_CLIENT_ID and AUTH_CLIENT_SECRET and AUTH_SERVER_METADATA_URL)


def _is_production() -> bool:
    """Detect production environment (HTTPS or explicit env)."""
    env = os.getenv("ENVIRONMENT", "").lower()
    if env == "prod" or env == "production":
        return True
    base_url = os.getenv("PUBLIC_BASE_URL", "")
    return base_url.lower().startswith("https://")


def _ensure_dirs() -> None:
    UPLOADS_DIR.mkdir(parents=True, exist_ok=True)


@asynccontextmanager
async def lifespan(app: FastAPI):
    log_level = os.getenv("LOG_LEVEL", "info").upper()
    logging.basicConfig(
        level=getattr(logging, log_level, logging.INFO),
        format="%(asctime)s %(levelname)s %(name)s %(message)s",
    )
    _ensure_dirs()
    init_db()

    # Ensure Databricks Delta tables exist (non-blocking on failure)
    try:
        await databricks.ensure_tables()
    except Exception as exc:
        logger.warning("Databricks table init failed (will retry on next submission): %s", exc)

    logger.info("Automower Trade-in Backend started")
    yield


app = FastAPI(title="Automower Trade-in Backend", lifespan=lifespan)

# Session middleware for SSO
session_secret = AUTH_COOKIE_SECRET or "change-me-session-secret-for-production"
https_only = _is_production()
if _is_sso_configured() and not AUTH_COOKIE_SECRET and https_only:
    logger.warning(
        "SSO configured but AUTH_COOKIE_SECRET is empty in production. "
        "Set AUTH_COOKIE_SECRET=|DOMAIN| and ensure Key Vault has the value."
    )
app.add_middleware(SessionMiddleware, secret_key=session_secret, same_site="lax", https_only=https_only)

oauth: Optional[OAuth] = None

if _is_sso_configured():
    oauth = OAuth()
    oauth.register(
        name="entra",
        client_id=AUTH_CLIENT_ID,
        client_secret=AUTH_CLIENT_SECRET,
        server_metadata_url=AUTH_SERVER_METADATA_URL,
        client_kwargs={"scope": "openid email profile"},
    )


def _healthz_payload() -> Dict[str, Any]:
    """Shared health check payload for DevPlatform (v2 format)."""
    build_id = os.getenv("BUILD_ID", "")
    return {
        "status": "ok",
        "buildInfo": {
            "buildId": build_id,
            "buildNumber": os.getenv("BUILD_NUMBER", ""),
            "sourceVersion": os.getenv("SOURCE_VERSION", ""),
            "serviceInstance": os.getenv("SERVICE_INSTANCE", ""),
            "serviceName": os.getenv("SERVICE_NAME", ""),
            "serviceDomain": os.getenv("SERVICE_DOMAIN", ""),
            "logLevel": os.getenv("LOG_LEVEL", "info"),
        },
        "info": {},
        "details": {},
    }


@app.get("/healthz")
async def healthz() -> JSONResponse:
    """Health check endpoint for DevPlatform pipeline (v2 format)."""
    return JSONResponse(_healthz_payload())


@app.get("/healthz/liveness")
async def healthz_liveness() -> Response:
    """Liveness probe: returns plain text 'UP'."""
    return Response(content="UP", media_type="text/plain")


@app.get("/healthz/readiness")
async def healthz_readiness() -> JSONResponse:
    """Readiness probe: same payload as /healthz."""
    return JSONResponse(_healthz_payload())


def _build_public_base_url(request: Request) -> str:
    public_base_url = os.getenv("PUBLIC_BASE_URL")
    if public_base_url:
        return public_base_url.rstrip("/")
    return str(request.base_url).rstrip("/")


def _build_file_url(path: Optional[str], request: Request) -> Optional[str]:
    if not path:
        return None
    if path.startswith("http://") or path.startswith("https://"):
        return path
    base_url = _build_public_base_url(request)
    return f"{base_url}/api/files/{path.lstrip('/')}"


async def notify_n8n_new_submission(submission_id: int, payload: Dict[str, Any]) -> None:
    webhook_url = os.getenv("N8N_WEBHOOK_URL")
    if not webhook_url:
        return
    try:
        async with httpx.AsyncClient(timeout=5.0) as client:
            await client.post(
                webhook_url,
                json={"submissionId": submission_id, "payload": payload},
            )
    except Exception as e:
        logger.warning("n8n webhook failed: %s", e)


def _is_admin_request(request: Request) -> bool:
    user = request.session.get("user") if hasattr(request, "session") else None
    return bool(user and isinstance(user, dict) and user.get("email"))


def get_current_admin(request: Request) -> str:
    if not _is_admin_request(request):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Unauthorized")
    return "admin"


app.mount("/static", StaticFiles(directory=str(BASE_DIR), html=False), name="static")
app.mount("/uploads", StaticFiles(directory=str(UPLOADS_DIR), html=False), name="uploads")

SUPPORTED_LANG_PATHS = {"en", "de-at"}


@app.get("/")
async def index_root() -> RedirectResponse:
    return RedirectResponse(url="/en", status_code=status.HTTP_302_FOUND)


@app.get("/en")
@app.get("/de-at")
async def index_localized() -> FileResponse:
    index_path = BASE_DIR / "index.html"
    if not index_path.exists():
        raise HTTPException(status_code=404, detail="index.html not found")
    return FileResponse(str(index_path))


@app.get("/admin-login")
async def admin_login_page(request: Request) -> Response:
    if oauth is None:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="SSO is not configured.")
    redirect_uri = AUTH_REDIRECT_URI or str(request.url_for("auth_callback"))
    return await oauth.entra.authorize_redirect(request, redirect_uri)


@app.get("/admin")
async def admin_page(request: Request) -> Response:
    if not _is_admin_request(request):
        return RedirectResponse(url="/admin-login")
    admin_path = BASE_DIR / "admin.html"
    if not admin_path.exists():
        raise HTTPException(status_code=404, detail="admin.html not found")
    return FileResponse(str(admin_path))


@app.get("/oauth2callback", name="auth_callback")
async def auth_callback(request: Request) -> Response:
    if oauth is None:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="SSO is not configured.")
    token = await oauth.entra.authorize_access_token(request)
    userinfo: Dict[str, Any] = token.get("userinfo") or {}
    if not userinfo:
        userinfo = token.get("id_token_claims", {})
    email = userinfo.get("email") or userinfo.get("preferred_username")
    name = userinfo.get("name") or email or "Admin"
    request.session["user"] = {"email": email, "name": name, "sub": userinfo.get("sub")}
    return RedirectResponse(url="/admin")


@app.get("/admin/logout")
async def admin_sso_logout(request: Request) -> Response:
    request.session.pop("user", None)
    return RedirectResponse(url="/")


def _save_upload_file(
    key: str,
    upload,
    submission_id: int,
    product_index: int,
) -> str:
    """Save uploaded file locally (temp buffer) and return its relative path."""
    if not upload or not getattr(upload, "filename", None):
        return ""

    suffix = (Path(upload.filename).suffix or "").lower()
    if suffix not in ALLOWED_UPLOAD_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"File type not allowed. Allowed: {', '.join(ALLOWED_UPLOAD_EXTENSIONS)}",
        )

    content = upload.file.read()
    if len(content) > MAX_UPLOAD_SIZE_BYTES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"File too large. Max size: {MAX_UPLOAD_SIZE_BYTES // (1024 * 1024)} MB",
        )
    upload.file.seek(0)

    safe_key = "".join(c if c.isalnum() or c in "-_" else "_" for c in key)
    filename = f"{submission_id}_{product_index}_{safe_key}{suffix}"
    destination = UPLOADS_DIR / filename

    with destination.open("wb") as buffer:
        buffer.write(content)

    return f"uploads/{filename}"


@app.post("/api/submissions")
async def create_submission(request: Request, background_tasks: BackgroundTasks) -> JSONResponse:
    """
    Accept form submissions. Saves to SQLite immediately, then syncs to Databricks
    (file upload to Unity Catalog Volume + Delta table insert) in the background.
    """
    form = await request.form()

    payload_raw = form.get("payload")
    if not payload_raw:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Missing payload field")

    try:
        payload: Dict[str, Any] = json.loads(payload_raw)
    except json.JSONDecodeError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid JSON in payload field")

    language = payload.get("language") or "en"
    dealer = payload.get("dealer") or {}
    base_product = payload.get("product") or {}
    additional_products = payload.get("additionalProducts") or []
    submitted_at = payload.get("submittedAt") or datetime.now(tz=timezone.utc).isoformat()

    file_objects: Dict[str, Any] = {}
    for key, value in form.multi_items():
        if hasattr(value, "filename") and value.filename:
            file_objects[key] = value

    all_products: List[Dict[str, Any]] = []

    def build_product(product: Dict[str, Any], product_index: int) -> Dict[str, Any]:
        return {
            "product_index": product_index,
            "sold_model": product.get("soldModel") or product.get("sold_model") or "",
            "new_serial_number": product.get("newSerialNumber") or product.get("new_serial_number") or "",
            "trade_in_type": product.get("tradeInType") or product.get("trade_in_type") or "",
            "trade_in_serial_number": product.get("tradeInSerialNumber") or product.get("trade_in_serial_number") or None,
            "trade_in_nameplate_key": product.get("tradeInNameplateKey"),
            "trade_in_product_image_key": product.get("tradeInProductImageKey"),
            "invoice_key": product.get("invoiceKey"),
            "trade_in_nameplate_path": None,
            "trade_in_product_image_path": None,
            "invoice_path": None,
        }

    all_products.append(build_product(base_product, 0))
    for idx, prod in enumerate(additional_products, start=1):
        all_products.append(build_product(prod, idx))

    # 1. Insert into SQLite (fast, synchronous)
    submission_id = insert_submission(
        submitted_at=submitted_at,
        language=language,
        dealer=dealer,
        products=[
            {
                "product_index": p["product_index"],
                "sold_model": p["sold_model"],
                "new_serial_number": p["new_serial_number"],
                "trade_in_type": p["trade_in_type"],
                "trade_in_serial_number": p.get("trade_in_serial_number"),
                "trade_in_nameplate_path": None,
                "trade_in_product_image_path": None,
                "invoice_path": None,
            }
            for p in all_products
        ],
    )

    # 2. Save files locally and update SQLite with paths
    from .db import get_connection
    conn = get_connection()
    try:
        cursor = conn.cursor()
        rows = cursor.execute(
            "SELECT id, product_index FROM products WHERE submission_id = ? ORDER BY product_index ASC, id ASC",
            (submission_id,),
        ).fetchall()
        from_db_products = [{"db_id": row["id"], "product_index": row["product_index"]} for row in rows]

        for prod in all_products:
            matching = next(
                (p for p in from_db_products if p["product_index"] == prod["product_index"]),
                None,
            )
            if not matching:
                continue

            db_id = matching["db_id"]
            nameplate_path = ""
            product_image_path = ""
            invoice_path = ""

            if prod.get("trade_in_nameplate_key"):
                upload = file_objects.get(prod["trade_in_nameplate_key"])
                if upload:
                    nameplate_path = _save_upload_file(prod["trade_in_nameplate_key"], upload, submission_id, prod["product_index"])
            if prod.get("trade_in_product_image_key"):
                upload = file_objects.get(prod["trade_in_product_image_key"])
                if upload:
                    product_image_path = _save_upload_file(prod["trade_in_product_image_key"], upload, submission_id, prod["product_index"])
            if prod.get("invoice_key"):
                upload = file_objects.get(prod["invoice_key"])
                if upload:
                    invoice_path = _save_upload_file(prod["invoice_key"], upload, submission_id, prod["product_index"])

            cursor.execute(
                "UPDATE products SET trade_in_nameplate_path=?, trade_in_product_image_path=?, invoice_path=? WHERE id=?",
                (nameplate_path or None, product_image_path or None, invoice_path or None, db_id),
            )

        conn.commit()

        # Fetch final product state (with IDs and file paths) for Databricks sync
        final_rows = cursor.execute(
            """SELECT id, product_index, sold_model, new_serial_number, trade_in_type,
                      trade_in_serial_number, trade_in_nameplate_path,
                      trade_in_product_image_path, invoice_path
               FROM products WHERE submission_id = ? ORDER BY product_index ASC""",
            (submission_id,),
        ).fetchall()
        products_for_sync = [dict(row) for row in final_rows]

    finally:
        conn.close()

    # 3. Schedule Databricks sync in background (non-blocking)
    background_tasks.add_task(
        databricks.sync_submission,
        submission_id=submission_id,
        submitted_at=submitted_at,
        language=language,
        dealer=dealer,
        products=products_for_sync,
        uploads_dir=UPLOADS_DIR,
    )

    await notify_n8n_new_submission(submission_id, payload)

    return JSONResponse({"id": submission_id, "success": True})


@app.get("/api/submissions")
async def list_submissions(
    limit: int = Query(100, ge=1, le=10_000),
    offset: int = Query(0, ge=0),
    current_admin: str = Depends(get_current_admin),
) -> JSONResponse:
    """Return submissions from Databricks (falls back to SQLite if Databricks not configured)."""
    if databricks.is_configured():
        try:
            submissions = await databricks.fetch_submissions(limit=limit, offset=offset)
            return JSONResponse(submissions)
        except Exception as exc:
            logger.error("Databricks read failed, falling back to SQLite: %s", exc)

    submissions = fetch_submissions_sqlite(limit=limit, offset=offset)
    return JSONResponse(submissions)


@app.get("/api/admin/databricks-status")
async def databricks_status(current_admin: str = Depends(get_current_admin)) -> JSONResponse:
    """Debug endpoint: tests Databricks connectivity and returns status."""
    configured = databricks.is_configured()
    if not configured:
        return JSONResponse({
            "configured": False,
            "error": "DATABRICKS_HOST, DATABRICKS_TOKEN or DATABRICKS_WAREHOUSE_ID not set",
        })
    try:
        result = await databricks._run_sql("SELECT current_user(), current_timestamp()")
        _, rows = databricks._parse_rows(result)
        user = rows[0][0] if rows else "unknown"
        return JSONResponse({
            "configured": True,
            "databricks_user": user,
            "submissions_table": databricks._submissions_table(),
            "products_table": databricks._products_table(),
            "volume_prefix": databricks._volume_prefix(),
        })
    except Exception as exc:
        return JSONResponse({"configured": True, "error": str(exc)}, status_code=502)


@app.get("/api/files/{path:path}")
async def get_file(
    path: str,
    current_admin: str = Depends(get_current_admin),
) -> Response:
    """Proxy a file from Databricks Unity Catalog Volume to the admin browser."""
    if not databricks.is_configured():
        raise HTTPException(status_code=503, detail="Databricks file storage not configured")
    # Prevent path traversal – path must stay within the uploads volume
    normalised = os.path.normpath(path)
    if ".." in normalised or not normalised.startswith(databricks._VOLUME):
        raise HTTPException(status_code=400, detail="Invalid file path")
    try:
        content, content_type = await databricks.download_file(path)
    except httpx.HTTPStatusError as exc:
        if exc.response.status_code == 404:
            raise HTTPException(status_code=404, detail="File not found")
        raise HTTPException(status_code=502, detail="Failed to retrieve file from storage")
    except Exception as exc:
        logger.error("File proxy failed for %s: %s", path, exc)
        raise HTTPException(status_code=502, detail="Failed to retrieve file from storage")

    filename = Path(path).name
    return Response(
        content=content,
        media_type=content_type,
        headers={"Content-Disposition": f'inline; filename="{filename}"'},
    )


@app.get("/api/submissions/flat")
async def list_submissions_flat(
    request: Request,
    limit: int = Query(1000, ge=1, le=10_000),
    offset: int = Query(0, ge=0),
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    current_admin: str = Depends(get_current_admin),
) -> JSONResponse:
    """Flat product rows for analytics / Databricks ingestion."""
    if databricks.is_configured():
        try:
            submissions = await databricks.fetch_submissions(limit=limit, offset=offset)
        except Exception as exc:
            logger.error("Databricks read failed, falling back to SQLite: %s", exc)
            submissions = fetch_submissions_sqlite(limit=limit, offset=offset)
    else:
        submissions = fetch_submissions_sqlite(limit=limit, offset=offset)

    from_dt: Optional[datetime] = None
    to_dt: Optional[datetime] = None
    if from_date:
        from_dt = datetime.fromisoformat(from_date)
    if to_date:
        to_dt = datetime.fromisoformat(to_date)

    rows: List[Dict[str, Any]] = []
    for sub in submissions:
        submitted_at_str = sub["submitted_at"]
        try:
            submitted_at = datetime.fromisoformat(submitted_at_str)
        except Exception:
            submitted_at = None

        if submitted_at and from_dt and submitted_at < from_dt:
            continue
        if submitted_at and to_dt and submitted_at > to_dt:
            continue

        for product in sub["products"]:
            rows.append({
                "submissionId":         sub["id"],
                "submittedAt":          submitted_at_str,
                "language":             sub["language"],
                "dealerNo":             sub["dealer"]["dealerNo"],
                "companyName":          sub["dealer"]["companyName"],
                "postalLocation":       sub["dealer"]["postalLocation"],
                "email":                sub["dealer"]["email"],
                "productIndex":         product["productIndex"],
                "soldModel":            product["soldModel"],
                "newSerialNumber":      product["newSerialNumber"],
                "tradeInType":          product["tradeInType"],
                "tradeInSerialNumber":  product["tradeInSerialNumber"],
                "tradeInNameplateUrl":  _build_file_url(product.get("tradeInNameplatePath"), request),
                "tradeInProductImageUrl": _build_file_url(product.get("tradeInProductImagePath"), request),
                "invoiceUrl":           _build_file_url(product.get("invoicePath"), request),
            })

    return JSONResponse(rows)


@app.get("/api/submissions/export")
async def export_submissions(
    request: Request,
    current_admin: str = Depends(get_current_admin),
) -> Response:
    """Export all submissions as Excel, with hyperlinks to files via /api/files/."""
    from io import BytesIO
    from openpyxl import Workbook

    if databricks.is_configured():
        try:
            submissions = await databricks.fetch_submissions(limit=10_000, offset=0)
        except Exception as exc:
            logger.error("Databricks read failed for export, falling back to SQLite: %s", exc)
            submissions = fetch_submissions_sqlite(limit=10_000, offset=0)
    else:
        submissions = fetch_submissions_sqlite(limit=10_000, offset=0)

    base_url = _build_public_base_url(request)

    def build_url(path: Optional[str]) -> Optional[str]:
        if not path:
            return None
        if path.startswith("http://") or path.startswith("https://"):
            return path
        return f"{base_url}/api/files/{path.lstrip('/')}"

    wb = Workbook()
    ws = wb.active
    ws.title = "Submissions"

    headers = [
        "submissionId", "submittedAt", "language", "dealerNo", "companyName",
        "postalLocation", "email", "productIndex", "soldModel", "newSerialNumber",
        "tradeInType", "tradeInSerialNumber",
        "tradeInNameplateLink", "tradeInProductImageLink", "invoiceLink",
    ]
    ws.append(headers)

    nameplate_col     = headers.index("tradeInNameplateLink") + 1
    product_image_col = headers.index("tradeInProductImageLink") + 1
    invoice_col       = headers.index("invoiceLink") + 1

    for sub in submissions:
        for product in sub["products"]:
            ws.append([
                sub["id"], sub["submitted_at"], sub["language"],
                sub["dealer"]["dealerNo"], sub["dealer"]["companyName"],
                sub["dealer"]["postalLocation"], sub["dealer"]["email"],
                product["productIndex"], product["soldModel"],
                product["newSerialNumber"], product["tradeInType"],
                product["tradeInSerialNumber"],
                "", "", "",
            ])
            row_idx = ws.max_row

            for col_idx, path_key in [
                (nameplate_col,     "tradeInNameplatePath"),
                (product_image_col, "tradeInProductImagePath"),
                (invoice_col,       "invoicePath"),
            ]:
                url = build_url(product.get(path_key))
                if url:
                    label = {"tradeInNameplatePath": "Nameplate", "tradeInProductImagePath": "Product image", "invoicePath": "Invoice"}[path_key]
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = label
                    cell.hyperlink = url
                    cell.style = "Hyperlink"

    for column_cells in ws.columns:
        max_length = max((len(str(cell.value)) for cell in column_cells if cell.value), default=0)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 40)

    buffer = BytesIO()
    wb.save(buffer)

    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="tradein_submissions.xlsx"'},
    )
